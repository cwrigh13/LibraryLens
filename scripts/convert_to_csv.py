#!/usr/bin/env python3
from __future__ import annotations
import csv
import io
import json
import pathlib
import re
import shutil
import sys
import tempfile
import zipfile
from typing import Dict, Any, List, Iterable

REPO_ROOT = pathlib.Path(__file__).resolve().parents[1]
RAW_ROOT = REPO_ROOT / 'data' / 'raw'
CSV_ROOT = REPO_ROOT / 'data' / 'csv'

try:
    import openpyxl  # type: ignore
except Exception:
    openpyxl = None


def ensure_parent(path: pathlib.Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def write_csv(rows: Iterable[Dict[str, Any]], fieldnames: List[str], dest: pathlib.Path) -> int:
    ensure_parent(dest)
    count = 0
    with dest.open('w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        for row in rows:
            writer.writerow({k: _stringify(row.get(k)) for k in fieldnames})
            count += 1
    return count


def _stringify(value: Any) -> str:
    if value is None:
        return ''
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def convert_csv_like(src: pathlib.Path, dest: pathlib.Path, delimiter: str = ',') -> int:
    ensure_parent(dest)
    with src.open('r', encoding='utf-8', errors='ignore', newline='') as f_in, dest.open('w', encoding='utf-8', newline='') as f_out:
        reader = csv.reader(f_in, delimiter=delimiter)
        writer = csv.writer(f_out)
        count = 0
        for row in reader:
            writer.writerow(row)
            count += 1
    return count


def convert_xlsx(src: pathlib.Path, dest_dir: pathlib.Path) -> List[pathlib.Path]:
    outputs: List[pathlib.Path] = []
    if openpyxl is None:
        return outputs
    wb = openpyxl.load_workbook(filename=str(src), read_only=True, data_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        safe_sheet = re.sub(r'[^A-Za-z0-9_-]+', '-', sheet).strip('-') or 'sheet'
        out = dest_dir / f"{src.stem}-{safe_sheet}.csv"
        ensure_parent(out)
        with out.open('w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow([_stringify(v) for v in row])
        outputs.append(out)
    return outputs


def flatten_geojson(obj: Dict[str, Any]) -> (List[Dict[str, Any]], List[str]):
    feats = obj.get('features') or []
    rows: List[Dict[str, Any]] = []
    # collect fieldnames from properties
    keys: List[str] = []
    key_set = set()
    for f in feats:
        props = f.get('properties') or {}
        for k in props.keys():
            if k not in key_set:
                key_set.add(k)
                keys.append(k)
    fieldnames = ['geometry_type', 'geometry'] + keys
    for f in feats:
        geom = f.get('geometry') or {}
        props = f.get('properties') or {}
        row = {'geometry_type': geom.get('type') or '', 'geometry': json.dumps(geom)}
        for k in keys:
            row[k] = props.get(k)
        rows.append(row)
    return rows, fieldnames


def convert_json(src: pathlib.Path, dest: pathlib.Path) -> List[pathlib.Path]:
    text = src.read_text(encoding='utf-8', errors='ignore')
    try:
        obj = json.loads(text)
    except Exception:
        return []
    outputs: List[pathlib.Path] = []
    # GeoJSON
    if isinstance(obj, dict) and obj.get('type') == 'FeatureCollection' and 'features' in obj:
        rows, fieldnames = flatten_geojson(obj)
        ensure_parent(dest)
        write_csv(rows, fieldnames, dest)
        outputs.append(dest)
        return outputs
    # Array of dicts
    if isinstance(obj, list) and obj and isinstance(obj[0], dict):
        # union keys
        keys: List[str] = []
        key_set = set()
        for r in obj:
            for k in r.keys():
                if k not in key_set:
                    key_set.add(k)
                    keys.append(k)
        ensure_parent(dest)
        write_csv(obj, keys, dest)
        outputs.append(dest)
        return outputs
    # Dict of arrays
    if isinstance(obj, dict):
        items = obj.get('items')
        if isinstance(items, list) and items and isinstance(items[0], dict):
            keys: List[str] = []
            key_set = set()
            for r in items:
                for k in r.keys():
                    if k not in key_set:
                        key_set.add(k)
                        keys.append(k)
            ensure_parent(dest)
            write_csv(items, keys, dest)
            outputs.append(dest)
            return outputs
    return outputs


def convert_zip(src: pathlib.Path, dest_dir: pathlib.Path) -> List[pathlib.Path]:
    outputs: List[pathlib.Path] = []
    with zipfile.ZipFile(src, 'r') as z:
        with tempfile.TemporaryDirectory() as td:
            z.extractall(td)
            for p in pathlib.Path(td).rglob('*'):
                if p.is_dir():
                    continue
                ext = p.suffix.lower()
                if ext in {'.csv', '.tsv'}:
                    out = dest_dir / f"{src.stem}-{p.name if ext=='.csv' else p.stem + '.csv'}"
                    outputs.append(pathlib.Path(out))
                    convert_csv_like(p, out, delimiter='\t' if ext == '.tsv' else ',')
                elif ext in {'.xlsx'} and openpyxl is not None:
                    outputs.extend(convert_xlsx(p, dest_dir))
                elif ext in {'.json', '.geojson'}:
                    out = dest_dir / f"{src.stem}-{p.stem}.csv"
                    out_list = convert_json(p, out)
                    outputs.extend(out_list)
                # other formats (SHP, GDB) are skipped without GIS libs
    return outputs


def process_manifest(mpath: pathlib.Path) -> List[pathlib.Path]:
    outputs: List[pathlib.Path] = []
    m = json.loads(mpath.read_text(encoding='utf-8'))
    base = mpath.parent
    # derive csv output base path preserving domain/dataset structure
    rel = base.relative_to(RAW_ROOT)
    out_dir = CSV_ROOT / rel
    for fname in m.get('saved_files', []) or []:
        src = base / fname
        if not src.exists():
            continue
        ext = src.suffix.lower()
        if ext == '.csv':
            out = out_dir / src.name
            convert_csv_like(src, out, ',')
            outputs.append(out)
        elif ext == '.tsv':
            out = out_dir / (src.stem + '.csv')
            convert_csv_like(src, out, '\t')
            outputs.append(out)
        elif ext in {'.xlsx'}:
            outputs.extend(convert_xlsx(src, out_dir))
        elif ext in {'.json', '.geojson'}:
            out = out_dir / (src.stem + '.csv')
            outputs.extend(convert_json(src, out))
        elif ext in {'.zip'}:
            outputs.extend(convert_zip(src, out_dir))
        else:
            # skip unsupported (pdf, shp, gdb, xls)
            pass
    return outputs


def main() -> int:
    all_outputs: List[pathlib.Path] = []
    for mpath in RAW_ROOT.rglob('manifest.json'):
        outs = process_manifest(mpath)
        all_outputs.extend(outs)
    print(f"Converted {len(all_outputs)} CSV files under {CSV_ROOT}")
    return 0


if __name__ == '__main__':
    raise SystemExit(main())